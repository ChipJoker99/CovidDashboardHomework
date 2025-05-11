from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session
from datetime import date
from typing import List, Optional
import io # Per BytesIO
import logging

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from starlette.responses import StreamingResponse

from app.db.session import get_db
from app.crud.crud_region_data import regional_data as crud_regional_data

router = APIRouter()
logger = logging.getLogger(__name__)

@router.get(
    "/regions.xlsx",
    summary="Export regional COVID-19 data to XLSX",
    description="Exports COVID-19 case data aggregated by Italian region for a specific date (or latest) to an XLSX file.",
    response_class=StreamingResponse
)
async def export_regional_data_to_excel(
    report_date: Optional[date] = Query(None, description="Date in YYYY-MM-DD format. If not provided, defaults to latest available."),
    sort_by: Optional[str] = Query(None, description="Field to sort by (e.g., 'total_positive_cases', 'region_name')."),
    sort_order: Optional[str] = Query("desc", description="Sort order ('asc' or 'desc').", pattern="^(asc|desc)$"),
    db: Session = Depends(get_db)
):
    """
    Endpoint to export regional COVID-19 data to an XLSX file.
    - If `report_date` is not provided, data for the latest available day in the DB is used.
    - Data is sorted according to `sort_by` and `sort_order` parameters.
    """
    target_date_to_query: Optional[date] = None

    if report_date:
        target_date_to_query = report_date
        if not crud_regional_data.data_exists_for_date(db, submission_date=target_date_to_query):
            logger.warning(f"Export requested for date {target_date_to_query}, but no data exists in DB.")
            raise HTTPException(status_code=404, detail=f"No data available in the database for date {target_date_to_query} to export.")
    else:
        target_date_to_query = crud_regional_data.get_latest_submission_date(db)
        if not target_date_to_query:
            logger.warning("Export requested for latest data, but database is empty.")
            raise HTTPException(status_code=404, detail="No data available in the database to export.")

    logger.info(f"Exporting data for date: {target_date_to_query}, Sort by: {sort_by}, Sort order: {sort_order}")

    db_records = crud_regional_data.get_by_date(
        db,
        submission_date=target_date_to_query,
        sort_by=sort_by,
        sort_order=sort_order.lower() if sort_order else "desc"
    )

    if not db_records:
        logger.warning(f"No records found for date {target_date_to_query} during export, though date was expected to have data.")
        raise HTTPException(status_code=404, detail=f"No data found to export for date {target_date_to_query}.")

    wb = Workbook()
    ws = wb.active
    ws.title = f"Dati Regioni {target_date_to_query.strftime('%Y-%m-%d')}"

    headers = ["Regione", "Totale Casi Positivi", "Data Riferimento"]
    ws.append(headers)

    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = cell.font.copy(bold=True)

    for record in db_records:
        ws.append([
            record.region_name,
            record.total_positive_cases,
            record.submission_date.strftime("%Y-%m-%d")
        ])

    column_widths = {}
    for row_idx, row_cells in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False)):
        for col_idx, cell in enumerate(row_cells, 1):
            if cell.value:
                current_width = len(str(cell.value)) + 2 
                column_letter = get_column_letter(col_idx)
                if column_letter not in column_widths or current_width > column_widths[column_letter]:
                    column_widths[column_letter] = current_width
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width


    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    file_name_date_str = target_date_to_query.strftime("%Y%m%d")
    output_filename = f"covid_data_regioni_{file_name_date_str}.xlsx"

    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={output_filename}"}
    )